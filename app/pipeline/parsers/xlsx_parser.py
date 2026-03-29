import re

from openpyxl import load_workbook

from .base import BaseParser, ContentBlock, ParsedContent

# Short alphanumeric codes (ABT-001, KST-4100, P-10284) are not PII
_CODE_PATTERN = re.compile(r"^[A-Z]{1,4}[-]?\d{2,6}$")


class XlsxParser(BaseParser):
    def parse(self, file_path: str) -> ParsedContent:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        blocks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    if cell.value is None:
                        continue

                    value = cell.value
                    text = str(value)

                    # Determine if this cell should skip PII detection
                    skip = False

                    # Row 0 = header row — never pseudonymize
                    if row_idx == 0:
                        skip = True

                    # Numeric values (int, float) — never pseudonymize
                    elif isinstance(value, (int, float)):
                        skip = True

                    # Short alphanumeric codes (ABT-001, KST-4100, P-10284)
                    elif _CODE_PATTERN.match(text.strip()):
                        skip = True

                    # Empty or whitespace-only
                    elif not text.strip():
                        continue

                    blocks.append(
                        ContentBlock(
                            text=text,
                            block_type="cell",
                            metadata={
                                "sheet": sheet_name,
                                "row": row_idx,
                                "col": col_idx,
                                "skip_detection": skip,
                            },
                        )
                    )

        wb.close()
        return ParsedContent(blocks=blocks, format="xlsx")
